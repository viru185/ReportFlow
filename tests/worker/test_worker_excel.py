"""End-to-end worker tests that drive a real Excel instance.

Marked ``excel`` so CI (which has no Excel) skips them via ``-m "not excel"``. Run locally
with ``uv run pytest -m excel``.

Every test asserts no net-new ``EXCEL.EXE`` process survives — the ghost check is the whole
point of the worker's teardown design.
"""

from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

import openpyxl
import psutil
import pytest

from reportflow.core.ipc import RunStatus, WorkerRequest, read_result, write_request
from reportflow.worker.runner import run_job

pytestmark = pytest.mark.excel


def _excel_pids() -> set[int]:
    pids: set[int] = set()
    for p in psutil.process_iter(["name"]):
        try:
            if (p.info["name"] or "").lower() == "excel.exe":
                pids.add(p.pid)
        except psutil.Error:
            pass
    return pids


def _make_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    data = wb.active
    data.title = "Data"
    for i, (u, pr) in enumerate([(120, 9.5), (80, 11.0), (200, 8.25), (150, 10.0)], start=2):
        data[f"B{i}"] = u
        data[f"C{i}"] = pr
        data[f"D{i}"] = f"=B{i}*C{i}"
    summary = wb.create_sheet("Summary")
    summary["B1"] = "=SUM(Data!B2:B5)"
    summary["B2"] = "=SUM(Data!D2:D5)"
    summary["B3"] = "=B2/B1"
    summary.print_area = "A1:B3"
    detail = wb.create_sheet("Detail")
    detail["B2"] = "=Summary!B2*2"
    detail["B3"] = "=Summary!B1+10"
    detail.print_area = "A1:B3"
    wb.save(path)


def _request(tmp_path: Path, sheets, **over) -> WorkerRequest:
    wb = tmp_path / "template.xlsx"
    if not wb.exists():
        _make_workbook(wb)
    defaults = dict(
        run_id="r1",
        job_name="j",
        input_excel_path=wb,
        output_xlsx_path=tmp_path / "out.xlsx",
        output_pdf_path=tmp_path / "{sheet}.pdf",
        sheet_names=sheets,
        timeout_seconds=120,
        is_test=True,
        result_path=tmp_path / "result.json",
        log_path=tmp_path / "worker.log",
    )
    defaults.update(over)
    return WorkerRequest(**defaults)


def test_success_freezes_and_exports(tmp_path):
    before = _excel_pids()
    result = run_job(_request(tmp_path, ["Summary", "Detail"]))

    assert result.status is RunStatus.SUCCESS
    assert Path(result.output_xlsx).exists()
    assert len(result.pdf_paths) == 2
    assert all(Path(p).stat().st_size > 0 for p in result.pdf_paths)
    assert result.excel_pid_reaped is True
    assert not (_excel_pids() - before), "ghost EXCEL.EXE leaked"

    wb = openpyxl.load_workbook(result.output_xlsx)
    assert wb["Summary"]["B1"].value == 550  # frozen to a value, not a formula
    assert str(wb["Data"]["D2"].value).startswith("=")  # non-selected sheet untouched


def test_missing_sheet_fails_cleanly(tmp_path):
    before = _excel_pids()
    result = run_job(_request(tmp_path, ["Summary", "DoesNotExist"]))

    assert result.status is RunStatus.FAILED
    assert "not found" in result.message.lower()
    assert result.excel_pid_reaped is True
    assert not (_excel_pids() - before)
    assert Path(tmp_path / "result.json").exists()


def test_missing_template_fails_cleanly(tmp_path):
    before = _excel_pids()
    req = _request(tmp_path, ["Summary"], input_excel_path=tmp_path / "nope.xlsx")
    result = run_job(req)

    assert result.status is RunStatus.FAILED
    assert result.excel_pid_reaped is True
    assert not (_excel_pids() - before)


def test_no_pdf_no_freeze_keeps_formulas(tmp_path):
    before = _excel_pids()
    req = _request(
        tmp_path, ["Summary"], generate_pdf=False, output_pdf_path=None, freeze_values=False
    )
    result = run_job(req)

    assert result.status is RunStatus.SUCCESS
    assert result.pdf_paths == []
    assert not (_excel_pids() - before)
    wb = openpyxl.load_workbook(result.output_xlsx)
    assert str(wb["Summary"]["B1"].value).startswith("=")  # not frozen -> still a formula


def test_parallel_subprocesses_no_ghost(tmp_path):
    """The real concurrency model: N separate worker processes at once (COM in a frozen
    subprocess), none leaking Excel."""
    before = _excel_pids()
    env = {**os.environ, "REPORTFLOW_DATA_DIR": str(tmp_path / "_data")}

    procs = []
    for i in range(3):
        run_dir = tmp_path / f"run{i}"
        run_dir.mkdir(parents=True, exist_ok=True)
        req = _request(
            tmp_path,
            ["Summary", "Detail"],
            run_id=f"run{i}",
            output_xlsx_path=run_dir / "out.xlsx",
            output_pdf_path=run_dir / "{sheet}.pdf",
            result_path=run_dir / "result.json",
            log_path=run_dir / "worker.log",
        )
        req_path = write_request(req, run_dir / "request.json")
        procs.append(
            (
                req,
                subprocess.Popen(
                    [sys.executable, "-m", "reportflow.worker", "--request", str(req_path)],
                    env=env,
                    creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                ),
            )
        )

    for req, proc in procs:
        assert proc.wait(timeout=180) == 0
        assert read_result(req.result_path).status is RunStatus.SUCCESS

    assert not (_excel_pids() - before), "ghost EXCEL.EXE leaked after parallel runs"
