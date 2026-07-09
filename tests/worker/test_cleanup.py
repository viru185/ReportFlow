"""Pure worker tests: output cleanup (openpyxl) and the empty-detection helper."""

from __future__ import annotations

import openpyxl

from reportflow.worker.cleanup import blank_out_values
from reportflow.worker.excel import _has_any_value


def _make_output(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Header"
    ws["A2"] = "Tag not found"
    ws["B2"] = 42
    ws["A3"] = "  No Data  "  # whitespace-padded junk still matches
    ws["B3"] = "#REF!"
    ws["C3"] = "keep me"
    wb.save(path)


def test_blank_out_values_removes_only_targets(tmp_path):
    out = tmp_path / "out.xlsx"
    _make_output(out)

    blanked = blank_out_values(out, ["Tag not found", "No Data", "#REF!"])
    assert blanked == 3

    wb = openpyxl.load_workbook(out)
    ws = wb["Report"]
    assert ws["A1"].value == "Header"
    assert ws["A2"].value is None
    assert ws["B2"].value == 42
    assert ws["A3"].value is None
    assert ws["B3"].value is None
    assert ws["C3"].value == "keep me"


def test_blank_out_values_noop_when_empty_list(tmp_path):
    out = tmp_path / "out.xlsx"
    _make_output(out)
    assert blank_out_values(out, []) == 0
    assert blank_out_values(out, ["  ", ""]) == 0
    wb = openpyxl.load_workbook(out)
    assert wb["Report"]["A2"].value == "Tag not found"  # untouched


def test_has_any_value_detects_data_and_emptiness():
    assert _has_any_value(None) is False
    assert _has_any_value("") is False
    assert _has_any_value("   ") is False
    assert _has_any_value([[None, None], [None, ""]]) is False
    assert _has_any_value([[None, None], [None, 0]]) is True  # zero is real data
    assert _has_any_value("text") is True
    assert _has_any_value([[None], ["x"]]) is True
    assert _has_any_value(3.14) is True
