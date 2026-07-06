"""End-to-end: the Service launches the REAL Excel worker and produces artifacts.

Marked ``excel`` (skipped in CI). Proves the full control-plane -> worker chain.
"""

from __future__ import annotations

import time
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from reportflow.core.config.loader import save_config
from reportflow.core.config.models import AppConfig, SmtpConfig
from reportflow.service.api import ServiceState, create_app

pytestmark = pytest.mark.excel


def _make_wb(path: Path) -> None:
    wb = openpyxl.Workbook()
    data = wb.active
    data.title = "Data"
    data["A1"] = 10
    data["A2"] = 20
    summary = wb.create_sheet("Summary")
    summary["A1"] = "=SUM(Data!A1:A2)"
    summary.print_area = "A1:A1"
    detail = wb.create_sheet("Detail")
    detail["A1"] = "=Summary!A1*2"
    detail.print_area = "A1:A1"
    wb.save(path)


def test_service_runs_real_worker(tmp_path):
    # SMTP to a refused port so the (real) run's email attempt fails fast and harmlessly.
    save_config(AppConfig(smtp=SmtpConfig(host="127.0.0.1", port=1, username="")))

    wb = tmp_path / "t.xlsx"
    _make_wb(wb)

    state = ServiceState()  # real worker (module command)
    app = create_app(state)
    with TestClient(app) as c:
        job = {
            "name": "e2e",
            "input_excel_path": str(wb),
            "output_dir": str(tmp_path / "out"),
            "output_name": "{run_id}",
            "sheet_names": ["Summary", "Detail"],
            "send_report_email": False,
            "prod": {"to": ["boss@corp.example.com"]},
            "test": {"to": ["dev@corp.example.com"]},
        }
        assert c.post("/jobs", json=job).status_code == 201

        run_id = c.post("/jobs/e2e/run").json()["run_id"]
        rec = None
        for _ in range(150):  # up to ~30s for a real Excel launch
            rec = c.get(f"/runs/{run_id}").json()
            if rec["status"] != "running":
                break
            time.sleep(0.2)

    assert rec is not None and rec["status"] == "success", rec
    out_xlsx = Path(rec["output_xlsx"])
    assert out_xlsx.exists()
    assert out_xlsx.parent == tmp_path / "out"  # honored output_dir
    assert len(rec["pdf_paths"]) == 2
    assert all(Path(p).exists() for p in rec["pdf_paths"])


def _wait_done(c: TestClient, run_id: str) -> dict:
    rec: dict = {}
    for _ in range(150):
        rec = c.get(f"/runs/{run_id}").json()
        if rec["status"] != "running":
            break
        time.sleep(0.2)
    return rec


def test_default_output_lands_next_to_input(tmp_path):
    save_config(AppConfig(smtp=SmtpConfig(host="127.0.0.1", port=1, username="")))
    wb = tmp_path / "input" / "t.xlsx"
    wb.parent.mkdir(parents=True)
    _make_wb(wb)

    state = ServiceState()
    app = create_app(state)
    with TestClient(app) as c:
        job = {
            "name": "e2e_default",
            "input_excel_path": str(wb),
            # no output_dir / output_name -> next to input, stem {job}_{date}
            "sheet_names": ["Summary"],
            "send_report_email": False,
            "prod": {"to": ["boss@corp.example.com"]},
            "test": {"to": ["dev@corp.example.com"]},
        }
        assert c.post("/jobs", json=job).status_code == 201
        rec = _wait_done(c, c.post("/jobs/e2e_default/run").json()["run_id"])

    assert rec["status"] == "success", rec
    out_xlsx = Path(rec["output_xlsx"])
    assert out_xlsx.parent == wb.parent  # defaulted next to the input file
    assert out_xlsx.name.startswith("e2e_default_")  # {job}_{date} stem
    assert len(rec["pdf_paths"]) == 1
    assert Path(rec["pdf_paths"][0]).name.endswith("_Summary.pdf")
