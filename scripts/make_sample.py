"""Generate a realistic sample workbook for exercising the Excel worker.

Creates ``scripts/sample/template.xlsx`` with:

* ``Data``   — raw numbers (a helper sheet that is NOT selected by the job).
* ``Summary``— formulas referencing ``Data``; landscape, fit-to-page print settings.
* ``Detail`` — formulas referencing ``Summary``; portrait print settings.

Run with:  uv run python scripts/make_sample.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins

OUT = Path(__file__).resolve().parent / "sample" / "template.xlsx"


def build() -> Path:
    wb = Workbook()

    data = wb.active
    data.title = "Data"
    data["A1"] = "Region"
    data["B1"] = "Units"
    data["C1"] = "Price"
    rows = [
        ("North", 120, 9.5),
        ("South", 80, 11.0),
        ("East", 200, 8.25),
        ("West", 150, 10.0),
    ]
    for i, (region, units, price) in enumerate(rows, start=2):
        data[f"A{i}"] = region
        data[f"B{i}"] = units
        data[f"C{i}"] = price
        data[f"D{i}"] = f"=B{i}*C{i}"  # formula: revenue per region

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Total Units"
    summary["B1"] = "=SUM(Data!B2:B5)"
    summary["A2"] = "Total Revenue"
    summary["B2"] = "=SUM(Data!D2:D5)"
    summary["A3"] = "Avg Price"
    summary["B3"] = "=B2/B1"
    # Print settings owned by the template.
    summary.page_setup.orientation = "landscape"
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.page_setup.fitToWidth = 1
    summary.page_setup.fitToHeight = 1
    summary.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    summary.print_area = "A1:B3"

    detail = wb.create_sheet("Detail")
    detail["A1"] = "Metric"
    detail["B1"] = "Value"
    detail["A2"] = "Revenue x2"
    detail["B2"] = "=Summary!B2*2"
    detail["A3"] = "Units + 10"
    detail["B3"] = "=Summary!B1+10"
    detail.page_setup.orientation = "portrait"
    detail.print_area = "A1:B3"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote sample workbook: {path}")
