"""Post-save output cleanup (openpyxl — no Excel/COM needed, unit-testable).

Blanks out configured junk values (PI DataLink error strings like "Tag not found",
"#REF!", …) from the OUTPUT workbook, mirroring the customer's legacy clean-up step.
Operates only on the saved output file, never the source.
"""

from __future__ import annotations

from pathlib import Path

from loguru import logger
from openpyxl import load_workbook


def blank_out_values(output_xlsx: Path, values: list[str]) -> int:
    """Replace any cell whose text equals one of ``values`` with an empty cell.

    Returns the number of cells blanked. Comparison is against the cell's string form,
    stripped — matching how the junk strings appear in practice.
    """
    targets = {v.strip() for v in values if v.strip()}
    if not targets:
        return 0

    workbook = load_workbook(output_xlsx)
    blanked = 0
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    if str(cell.value).strip() in targets:
                        cell.value = None
                        blanked += 1
        if blanked:
            workbook.save(output_xlsx)
    finally:
        workbook.close()
    logger.info("Blanked {} junk cell(s) in {}", blanked, output_xlsx.name)
    return blanked
