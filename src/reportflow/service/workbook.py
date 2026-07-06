"""Workbook sheet discovery for the UI — via openpyxl, never COM.

Keeping Excel out of the Service is a hard rule; sheet names are read statically.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


class WorkbookError(Exception):
    pass


def discover_sheets(path: Path) -> list[str]:
    p = Path(path)
    if not p.exists():
        raise WorkbookError(f"workbook not found: {p}")
    try:
        wb = load_workbook(p, read_only=True, keep_links=False)
    except Exception as e:  # noqa: BLE001 — surface a clean message to the UI
        raise WorkbookError(f"could not open workbook: {e}") from e
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()
